from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from automation.reporting import (
    build_person_attributes_enhanced_headers,
    build_person_attributes_enhanced_query,
    render_person_attributes_enhanced_workbook,
)


class PersonAttributesEnhancedReportTest(unittest.TestCase):
    def test_build_headers_and_query(self) -> None:
        person_columns = ["工号", "姓名", "部门ID"]

        headers = build_person_attributes_enhanced_headers(person_columns)
        query = build_person_attributes_enhanced_query(person_columns)

        self.assertEqual(headers, ["工号", "姓名", "部门ID", "所在城市", "组织单位", "所属战区"])
        self.assertIn('FROM "人员属性查询" AS p', query)
        self.assertIn('LEFT JOIN "组织属性查询" AS o', query)
        self.assertIn('p."部门ID"', query)
        self.assertIn('o."行政组织编码"', query)
        self.assertIn('o."所在城市" AS "所在城市"', query)
        self.assertIn('ORDER BY p."工号"', query)

    def test_render_workbook(self) -> None:
        headers = ["工号", "姓名", "所在城市", "组织单位", "所属战区"]
        rows = [
            ("1001", "张三", "深圳", "物业BG", "南部战区"),
            ("1002", "李四", None, "", None),
        ]

        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "person_attributes_enhanced.xlsx"
            row_count = render_person_attributes_enhanced_workbook(
                headers=headers,
                rows=rows,
                output_path=output_path,
            )

            self.assertEqual(row_count, 2)
            self.assertTrue(output_path.exists())

            workbook = load_workbook(output_path)
            sheet = workbook["人员属性查询增强报表"]
            exported_rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(exported_rows[0], tuple(headers))
        self.assertEqual(exported_rows[1], rows[0])
        self.assertEqual(exported_rows[2], ("1002", "李四", None, None, None))
