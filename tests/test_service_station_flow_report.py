from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from automation.reporting.service_station_flow import (
    build_service_station_flow_report,
    render_service_station_flow_workbook,
)
from automation.utils.war_zone import derive_war_zone_from_org_path


class ServiceStationFlowReportTest(unittest.TestCase):
    def test_derive_war_zone_from_org_path(self) -> None:
        self.assertEqual(
            derive_war_zone_from_org_path(
                "万物云_万物云本部_人力资源与行政服务中心_福州人力资源与行政服务部_福州招聘组"
            ),
            "福建战区",
        )
        self.assertEqual(
            derive_war_zone_from_org_path(
                "万物云_万物云本部_人力资源与行政服务中心_东北人力资源与行政服务部_沈阳东湖人事交付服务组"
            ),
            "东北战区",
        )

    def test_build_service_station_flow_report(self) -> None:
        start_rows = [
            {
                "employee_no": "1001",
                "employee_name": "甲",
                "department_id": "D-1",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海一组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": None,
            },
            {
                "employee_no": "1002",
                "employee_name": "乙",
                "department_id": "D-2",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海二组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1003",
                "employee_name": "丙",
                "department_id": "D-3",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海三组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1004",
                "employee_name": "丁",
                "department_id": "D-4",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "招聘",
                "standard_position_name": "招聘岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_东北人力资源与行政服务部_沈阳招聘组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-招聘",
                "war_zone": "东北战区",
            },
            {
                "employee_no": "1005",
                "employee_name": "戊",
                "department_id": "D-5",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_广州人力资源与行政服务部_广州一组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "广州战区",
            },
            {
                "employee_no": "1006",
                "employee_name": "己",
                "department_id": "D-6",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海四组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1007",
                "employee_name": "庚",
                "department_id": "D-7",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海五组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1008",
                "employee_name": "辛",
                "department_id": "D-8",
                "org_unit_name": "组织发展中心",
                "position_name": "HRBP",
                "standard_position_name": "HRBP岗",
                "org_path_name": "组织发展中心_上海HRBP组",
                "hr_type": "H1",
                "hr_subdomain": "HRBP/业务支持",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1009",
                "employee_name": "壬",
                "department_id": "D-9",
                "org_unit_name": "综合管理部",
                "position_name": "运营支持",
                "standard_position_name": "运营支持岗",
                "org_path_name": "综合管理部_广州组",
                "hr_type": "X1",
                "hr_subdomain": None,
                "war_zone": "广州战区",
            },
        ]
        end_rows = [
            {
                "employee_no": "1002",
                "employee_name": "乙",
                "department_id": "D-2",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海二组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1003",
                "employee_name": "丙",
                "department_id": "D-3",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "招聘",
                "standard_position_name": "招聘岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海招聘组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-招聘",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1004",
                "employee_name": "丁",
                "department_id": "D-4",
                "org_unit_name": "组织发展中心",
                "position_name": "HRBP",
                "standard_position_name": "HRBP岗",
                "org_path_name": "组织发展中心_东北HRBP组",
                "hr_type": "H1",
                "hr_subdomain": "HRBP/业务支持",
                "war_zone": "东北战区",
            },
            {
                "employee_no": "1005",
                "employee_name": "戊",
                "department_id": "D-5",
                "org_unit_name": "综合管理部",
                "position_name": "运营支持",
                "standard_position_name": "运营支持岗",
                "org_path_name": "综合管理部_广州组",
                "hr_type": "X1",
                "hr_subdomain": None,
                "war_zone": "广州战区",
            },
            {
                "employee_no": "1006",
                "employee_name": "己",
                "department_id": "D-10",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_苏州人力资源与行政服务部_苏州一组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "苏州战区",
            },
            {
                "employee_no": "1007",
                "employee_name": "庚",
                "department_id": "D-11",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "招聘",
                "standard_position_name": "招聘岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_苏州人力资源与行政服务部_苏州招聘组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-招聘",
                "war_zone": "苏州战区",
            },
            {
                "employee_no": "1008",
                "employee_name": "辛",
                "department_id": "D-12",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "招聘",
                "standard_position_name": "招聘岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海招聘二组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-招聘",
                "war_zone": "上海战区",
            },
            {
                "employee_no": "1009",
                "employee_name": "壬",
                "department_id": "D-13",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_广州人力资源与行政服务部_广州二组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "广州战区",
            },
            {
                "employee_no": "1010",
                "employee_name": "癸",
                "department_id": "D-14",
                "org_unit_name": "人力资源与行政服务中心",
                "position_name": "人事运营",
                "standard_position_name": "人事运营岗",
                "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_广州人力资源与行政服务部_广州三组",
                "hr_type": "H1",
                "hr_subdomain": "服务站-人事运营",
                "war_zone": "广州战区",
            },
        ]

        result = build_service_station_flow_report(
            start_date=date(2025, 12, 31),
            end_date=date(2026, 3, 30),
            start_rows=start_rows,
            end_rows=end_rows,
        )

        self.assertEqual(result["summary"]["startTargetCount"], 7)
        self.assertEqual(result["summary"]["endTargetCount"], 7)
        self.assertEqual(result["summary"]["leftCount"], 1)
        self.assertEqual(result["summary"]["otherHrOutCount"], 1)
        self.assertEqual(result["summary"]["nonHrOutCount"], 1)
        self.assertEqual(result["summary"]["targetSwitchCount"], 2)
        self.assertEqual(result["summary"]["sameSubdomainSameWarZoneCount"], 1)
        self.assertEqual(result["summary"]["sameSubdomainCrossWarZoneCount"], 1)
        self.assertEqual(result["summary"]["otherHrInCount"], 1)
        self.assertEqual(result["summary"]["nonHrInCount"], 1)
        self.assertEqual(result["summary"]["newInCount"], 1)

        shanghai_row = next(row for row in result["zoneSummaryRows"] if row["warZone"] == "上海战区")
        self.assertEqual(shanghai_row["leftCount"], 1)
        self.assertEqual(shanghai_row["opsToRecruitCount"], 2)
        self.assertEqual(len(result["leftRows"]), 1)
        self.assertEqual(len(result["otherHrOutRows"]), 1)
        self.assertEqual(len(result["targetFlowRows"]), 7)
        self.assertEqual(len(result["otherHrInRows"]), 1)
        self.assertEqual(len(result["allEndRows"]), 7)
        self.assertEqual(result["allEndRows"][0]["employeeNo"], "1002")
        self.assertEqual(result["allEndRows"][0]["endSubdomain"], "服务站-人事运营")

    def test_render_service_station_flow_workbook(self) -> None:
        report = build_service_station_flow_report(
            start_date=date(2025, 12, 31),
            end_date=date(2026, 3, 30),
            start_rows=[
                {
                    "employee_no": "1001",
                    "employee_name": "甲",
                    "department_id": "D-1",
                    "org_unit_name": "人力资源与行政服务中心",
                    "position_name": "人事运营",
                    "standard_position_name": "人事运营岗",
                    "org_path_name": "万物云_万物云本部_人力资源与行政服务中心_上海人力资源与行政服务部_上海一组",
                    "hr_type": "H1",
                    "hr_subdomain": "服务站-人事运营",
                    "war_zone": "上海战区",
                }
            ],
            end_rows=[],
        )

        with TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "service_station_flow.xlsx"
            render_service_station_flow_workbook(report=report, output_path=output_path)
            workbook = load_workbook(output_path)

        self.assertEqual(
            workbook.sheetnames,
            ["汇总", "战区汇总", "离职明细", "转其他HR明细", "目标岗位去向明细", "其他HR转入明细", "全部期末清单"],
        )
        self.assertEqual(workbook["离职明细"].max_row, 2)
        self.assertEqual(workbook["全部期末清单"].max_row, 1)


if __name__ == "__main__":
    unittest.main()
