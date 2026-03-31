from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from automation.utils.war_zone import resolve_person_war_zone

TARGET_SUBDOMAINS: tuple[str, str] = ("服务站-人事运营", "服务站-招聘")
TARGET_SUBDOMAIN_SET = set(TARGET_SUBDOMAINS)

OUTFLOW_CATEGORY_ORDER: tuple[str, ...] = (
    "留任原子域原战区",
    "留任原子域跨战区",
    "目标岗位内转子域(同战区)",
    "目标岗位内转子域并跨战区",
    "转其他HR岗位",
    "转非HR岗位",
    "离职",
)

INFLOW_CATEGORY_ORDER: tuple[str, ...] = (
    "来自原子域原战区",
    "来自原子域跨战区",
    "来自另一目标子域",
    "来自其他HR岗位",
    "来自非HR岗位",
    "期初不在目标岗位",
)

SUMMARY_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("metric", "指标"),
    ("value", "数值"),
)

ZONE_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("warZone", "战区"),
    ("startOperationsCount", "人事运营期初人数"),
    ("endOperationsCount", "人事运营期末人数"),
    ("operationsDelta", "人事运营增减"),
    ("startRecruitCount", "招聘期初人数"),
    ("endRecruitCount", "招聘期末人数"),
    ("recruitDelta", "招聘增减"),
    ("startTotalCount", "合计期初人数"),
    ("endTotalCount", "合计期末人数"),
    ("totalDelta", "合计增减"),
    ("leftCount", "离职人数"),
    ("otherHrOutCount", "转其他HR人数"),
    ("otherHrInCount", "其他HR转入人数"),
    ("opsToRecruitCount", "人事运营转招聘人数"),
    ("recruitToOpsCount", "招聘转人事运营人数"),
)

OUTFLOW_DETAIL_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("employeeNo", "工号"),
    ("employeeName", "姓名"),
    ("startSubdomain", "期初子域"),
    ("startWarZone", "期初战区"),
    ("startOrgUnitName", "期初组织单位"),
    ("startDepartmentId", "期初部门ID"),
    ("startPositionName", "期初职位名称"),
    ("startStandardPositionName", "期初标准岗位名称"),
    ("startHrType", "期初HR类型"),
    ("startOrgPathName", "期初组织路径名称"),
    ("movementType", "去向类型"),
    ("endSubdomain", "期末子域"),
    ("endWarZone", "期末战区"),
    ("endOrgUnitName", "期末组织单位"),
    ("endDepartmentId", "期末部门ID"),
    ("endPositionName", "期末职位名称"),
    ("endStandardPositionName", "期末标准岗位名称"),
    ("endHrType", "期末HR类型"),
    ("endOrgPathName", "期末组织路径名称"),
)

INFLOW_DETAIL_EXPORT_COLUMNS: tuple[tuple[str, str], ...] = (
    ("employeeNo", "工号"),
    ("employeeName", "姓名"),
    ("startSubdomain", "期初子域"),
    ("startWarZone", "期初战区"),
    ("startOrgUnitName", "期初组织单位"),
    ("startDepartmentId", "期初部门ID"),
    ("startPositionName", "期初职位名称"),
    ("startStandardPositionName", "期初标准岗位名称"),
    ("startHrType", "期初HR类型"),
    ("startOrgPathName", "期初组织路径名称"),
    ("movementType", "来源类型"),
    ("endSubdomain", "期末子域"),
    ("endWarZone", "期末战区"),
    ("endOrgUnitName", "期末组织单位"),
    ("endDepartmentId", "期末部门ID"),
    ("endPositionName", "期末职位名称"),
    ("endStandardPositionName", "期末标准岗位名称"),
    ("endHrType", "期末HR类型"),
    ("endOrgPathName", "期末组织路径名称"),
)


@dataclass(frozen=True)
class ServiceStationFlowWorkbookData:
    summary_rows: list[dict[str, Any]]
    zone_summary_rows: list[dict[str, Any]]
    left_rows: list[dict[str, Any]]
    other_hr_out_rows: list[dict[str, Any]]
    target_flow_rows: list[dict[str, Any]]
    other_hr_in_rows: list[dict[str, Any]]


def _strip_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    normalized = str(value).strip()
    return normalized or None


def _text_or_empty(value: Any) -> str:
    return _strip_text(value) or ""


def _text_or_dash(value: Any) -> str:
    return _strip_text(value) or "-"


def _is_hr_type(value: Any) -> bool:
    normalized = _strip_text(value)
    return bool(normalized and normalized.startswith("H"))


def _is_target_subdomain(value: Any) -> bool:
    normalized = _strip_text(value)
    return normalized in TARGET_SUBDOMAIN_SET


def _sort_employee_no(value: str) -> tuple[int, str]:
    normalized = value.strip()
    return (0, normalized) if normalized.isdigit() else (1, normalized)


def normalize_service_station_snapshot(row: dict[str, Any]) -> dict[str, Any]:
    employee_no = _strip_text(row.get("employee_no"))
    if employee_no is None:
        raise ValueError("service station snapshot row is missing employee_no")

    return {
        "employee_no": employee_no,
        "employee_name": _strip_text(row.get("employee_name")),
        "department_id": _strip_text(row.get("department_id")),
        "org_unit_name": _strip_text(row.get("org_unit_name")),
        "position_name": _strip_text(row.get("position_name")),
        "standard_position_name": _strip_text(row.get("standard_position_name")),
        "org_path_name": _strip_text(row.get("org_path_name")),
        "hr_type": _strip_text(row.get("hr_type")),
        "hr_subdomain": _strip_text(row.get("hr_subdomain")),
        "war_zone": resolve_person_war_zone(row.get("war_zone"), row.get("org_path_name")),
    }


def _build_detail_row(
    *,
    employee_no: str,
    employee_name: str | None,
    start_row: dict[str, Any] | None,
    end_row: dict[str, Any] | None,
    movement_type: str,
) -> dict[str, Any]:
    start_snapshot = start_row or {}
    end_snapshot = end_row or {}
    return {
        "id": employee_no,
        "employeeNo": employee_no,
        "employeeName": employee_name or _strip_text(start_snapshot.get("employee_name")) or _strip_text(end_snapshot.get("employee_name")) or "-",
        "startSubdomain": _text_or_dash(start_snapshot.get("hr_subdomain")),
        "startWarZone": _text_or_dash(start_snapshot.get("war_zone")),
        "startOrgUnitName": _text_or_dash(start_snapshot.get("org_unit_name")),
        "startDepartmentId": _text_or_dash(start_snapshot.get("department_id")),
        "startPositionName": _text_or_dash(start_snapshot.get("position_name")),
        "startStandardPositionName": _text_or_dash(start_snapshot.get("standard_position_name")),
        "startHrType": _text_or_dash(start_snapshot.get("hr_type")),
        "startOrgPathName": _text_or_dash(start_snapshot.get("org_path_name")),
        "movementType": movement_type,
        "endSubdomain": _text_or_dash(end_snapshot.get("hr_subdomain")),
        "endWarZone": _text_or_dash(end_snapshot.get("war_zone")),
        "endOrgUnitName": _text_or_dash(end_snapshot.get("org_unit_name")),
        "endDepartmentId": _text_or_dash(end_snapshot.get("department_id")),
        "endPositionName": _text_or_dash(end_snapshot.get("position_name")),
        "endStandardPositionName": _text_or_dash(end_snapshot.get("standard_position_name")),
        "endHrType": _text_or_dash(end_snapshot.get("hr_type")),
        "endOrgPathName": _text_or_dash(end_snapshot.get("org_path_name")),
    }


def _categorize_outflow(start_row: dict[str, Any], end_row: dict[str, Any] | None) -> str:
    if end_row is None:
        return "离职"

    start_subdomain = _strip_text(start_row.get("hr_subdomain"))
    end_subdomain = _strip_text(end_row.get("hr_subdomain"))
    start_war_zone = _strip_text(start_row.get("war_zone"))
    end_war_zone = _strip_text(end_row.get("war_zone"))
    same_subdomain = start_subdomain == end_subdomain
    same_war_zone = start_war_zone == end_war_zone

    if _is_target_subdomain(end_subdomain):
        if same_subdomain and same_war_zone:
            return "留任原子域原战区"
        if same_subdomain:
            return "留任原子域跨战区"
        if same_war_zone:
            return "目标岗位内转子域(同战区)"
        return "目标岗位内转子域并跨战区"

    if _is_hr_type(end_row.get("hr_type")):
        return "转其他HR岗位"
    return "转非HR岗位"


def _categorize_inflow(end_row: dict[str, Any], start_row: dict[str, Any] | None) -> str:
    if start_row is None:
        return "期初不在目标岗位"

    start_subdomain = _strip_text(start_row.get("hr_subdomain"))
    end_subdomain = _strip_text(end_row.get("hr_subdomain"))
    start_war_zone = _strip_text(start_row.get("war_zone"))
    end_war_zone = _strip_text(end_row.get("war_zone"))

    if _is_target_subdomain(start_subdomain):
        if start_subdomain == end_subdomain and start_war_zone == end_war_zone:
            return "来自原子域原战区"
        if start_subdomain == end_subdomain:
            return "来自原子域跨战区"
        return "来自另一目标子域"

    if _is_hr_type(start_row.get("hr_type")):
        return "来自其他HR岗位"
    return "来自非HR岗位"


def _counter_rows(counter: Counter[str], order: Iterable[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for category in order:
        rows.append(
            {
                "id": category,
                "category": category,
                "count": int(counter.get(category, 0)),
            }
        )
    return rows


def _zone_labels(rows: Iterable[dict[str, Any]]) -> list[str]:
    labels = {
        _strip_text(row.get("war_zone")) or "未识别战区"
        for row in rows
    }
    return sorted(labels)


def build_service_station_flow_report(
    *,
    start_date: date,
    end_date: date,
    start_rows: Iterable[dict[str, Any]],
    end_rows: Iterable[dict[str, Any]],
) -> dict[str, Any]:
    normalized_start_rows = [normalize_service_station_snapshot(row) for row in start_rows]
    normalized_end_rows = [normalize_service_station_snapshot(row) for row in end_rows]

    start_all_by_employee = {row["employee_no"]: row for row in normalized_start_rows}
    end_all_by_employee = {row["employee_no"]: row for row in normalized_end_rows}
    start_target_rows = [row for row in normalized_start_rows if _is_target_subdomain(row.get("hr_subdomain"))]
    end_target_rows = [row for row in normalized_end_rows if _is_target_subdomain(row.get("hr_subdomain"))]

    outflow_counter: Counter[str] = Counter()
    inflow_counter: Counter[str] = Counter()
    other_hr_out_destination_counter: Counter[str] = Counter()
    other_hr_in_source_counter: Counter[str] = Counter()

    target_flow_rows: list[dict[str, Any]] = []
    left_rows: list[dict[str, Any]] = []
    other_hr_out_rows: list[dict[str, Any]] = []
    other_hr_in_rows: list[dict[str, Any]] = []

    for start_row in start_target_rows:
        employee_no = start_row["employee_no"]
        end_row = end_all_by_employee.get(employee_no)
        movement_type = _categorize_outflow(start_row, end_row)
        outflow_counter[movement_type] += 1
        detail_row = _build_detail_row(
            employee_no=employee_no,
            employee_name=start_row.get("employee_name"),
            start_row=start_row,
            end_row=end_row,
            movement_type=movement_type,
        )
        target_flow_rows.append(detail_row)

        if movement_type == "离职":
            left_rows.append(detail_row)
        elif movement_type == "转其他HR岗位":
            other_hr_out_rows.append(detail_row)
            destination = _strip_text(end_row.get("hr_subdomain") if end_row else None) or "未识别其他HR子域"
            other_hr_out_destination_counter[destination] += 1

    for end_row in end_target_rows:
        employee_no = end_row["employee_no"]
        start_row = start_all_by_employee.get(employee_no)
        movement_type = _categorize_inflow(end_row, start_row)
        inflow_counter[movement_type] += 1
        if movement_type == "来自其他HR岗位":
            detail_row = _build_detail_row(
                employee_no=employee_no,
                employee_name=end_row.get("employee_name"),
                start_row=start_row,
                end_row=end_row,
                movement_type=movement_type,
            )
            other_hr_in_rows.append(detail_row)
            source = _strip_text(start_row.get("hr_subdomain") if start_row else None) or "未识别其他HR子域"
            other_hr_in_source_counter[source] += 1

    target_flow_rows.sort(key=lambda row: _sort_employee_no(row["employeeNo"]))
    left_rows.sort(key=lambda row: _sort_employee_no(row["employeeNo"]))
    other_hr_out_rows.sort(key=lambda row: _sort_employee_no(row["employeeNo"]))
    other_hr_in_rows.sort(key=lambda row: _sort_employee_no(row["employeeNo"]))

    zone_summary_rows: list[dict[str, Any]] = []
    zone_labels = sorted(set(_zone_labels(start_target_rows)) | set(_zone_labels(end_target_rows)))
    for zone_label in zone_labels:
        start_operations_count = sum(
            1
            for row in start_target_rows
            if (_strip_text(row.get("war_zone")) or "未识别战区") == zone_label and row.get("hr_subdomain") == "服务站-人事运营"
        )
        end_operations_count = sum(
            1
            for row in end_target_rows
            if (_strip_text(row.get("war_zone")) or "未识别战区") == zone_label and row.get("hr_subdomain") == "服务站-人事运营"
        )
        start_recruit_count = sum(
            1
            for row in start_target_rows
            if (_strip_text(row.get("war_zone")) or "未识别战区") == zone_label and row.get("hr_subdomain") == "服务站-招聘"
        )
        end_recruit_count = sum(
            1
            for row in end_target_rows
            if (_strip_text(row.get("war_zone")) or "未识别战区") == zone_label and row.get("hr_subdomain") == "服务站-招聘"
        )
        left_count = sum(
            1
            for row in left_rows
            if row["startWarZone"] == zone_label
        )
        other_hr_out_count = sum(
            1
            for row in other_hr_out_rows
            if row["startWarZone"] == zone_label
        )
        other_hr_in_count = sum(
            1
            for row in other_hr_in_rows
            if row["endWarZone"] == zone_label
        )
        ops_to_recruit_count = sum(
            1
            for row in target_flow_rows
            if row["startWarZone"] == zone_label
            and row["startSubdomain"] == "服务站-人事运营"
            and row["endSubdomain"] == "服务站-招聘"
            and row["movementType"] in {"目标岗位内转子域(同战区)", "目标岗位内转子域并跨战区"}
        )
        recruit_to_ops_count = sum(
            1
            for row in target_flow_rows
            if row["startWarZone"] == zone_label
            and row["startSubdomain"] == "服务站-招聘"
            and row["endSubdomain"] == "服务站-人事运营"
            and row["movementType"] in {"目标岗位内转子域(同战区)", "目标岗位内转子域并跨战区"}
        )
        start_total_count = start_operations_count + start_recruit_count
        end_total_count = end_operations_count + end_recruit_count
        zone_summary_rows.append(
            {
                "id": zone_label,
                "warZone": zone_label,
                "startOperationsCount": start_operations_count,
                "endOperationsCount": end_operations_count,
                "operationsDelta": end_operations_count - start_operations_count,
                "startRecruitCount": start_recruit_count,
                "endRecruitCount": end_recruit_count,
                "recruitDelta": end_recruit_count - start_recruit_count,
                "startTotalCount": start_total_count,
                "endTotalCount": end_total_count,
                "totalDelta": end_total_count - start_total_count,
                "leftCount": left_count,
                "otherHrOutCount": other_hr_out_count,
                "otherHrInCount": other_hr_in_count,
                "opsToRecruitCount": ops_to_recruit_count,
                "recruitToOpsCount": recruit_to_ops_count,
            }
        )

    summary = {
        "startTargetCount": len(start_target_rows),
        "endTargetCount": len(end_target_rows),
        "leftCount": int(outflow_counter.get("离职", 0)),
        "otherHrOutCount": int(outflow_counter.get("转其他HR岗位", 0)),
        "nonHrOutCount": int(outflow_counter.get("转非HR岗位", 0)),
        "targetSwitchCount": int(outflow_counter.get("目标岗位内转子域(同战区)", 0) + outflow_counter.get("目标岗位内转子域并跨战区", 0)),
        "sameSubdomainSameWarZoneCount": int(outflow_counter.get("留任原子域原战区", 0)),
        "sameSubdomainCrossWarZoneCount": int(outflow_counter.get("留任原子域跨战区", 0)),
        "otherHrInCount": int(inflow_counter.get("来自其他HR岗位", 0)),
        "nonHrInCount": int(inflow_counter.get("来自非HR岗位", 0)),
        "newInCount": int(inflow_counter.get("期初不在目标岗位", 0)),
    }

    summary_rows = [
        {"id": "startTargetCount", "metric": "期初目标岗位人数", "value": summary["startTargetCount"]},
        {"id": "endTargetCount", "metric": "期末目标岗位人数", "value": summary["endTargetCount"]},
        {"id": "leftCount", "metric": "离职人数", "value": summary["leftCount"]},
        {"id": "otherHrOutCount", "metric": "转其他HR人数", "value": summary["otherHrOutCount"]},
        {"id": "nonHrOutCount", "metric": "转非HR人数", "value": summary["nonHrOutCount"]},
        {"id": "targetSwitchCount", "metric": "目标岗位内互转人数", "value": summary["targetSwitchCount"]},
        {"id": "sameSubdomainSameWarZoneCount", "metric": "留任原子域原战区人数", "value": summary["sameSubdomainSameWarZoneCount"]},
        {"id": "sameSubdomainCrossWarZoneCount", "metric": "留任原子域跨战区人数", "value": summary["sameSubdomainCrossWarZoneCount"]},
        {"id": "otherHrInCount", "metric": "其他HR转入人数", "value": summary["otherHrInCount"]},
        {"id": "nonHrInCount", "metric": "非HR转入人数", "value": summary["nonHrInCount"]},
        {"id": "newInCount", "metric": "期初不在目标岗位人数", "value": summary["newInCount"]},
    ]

    other_hr_out_destination_rows = [
        {
            "id": category,
            "category": category,
            "count": count,
        }
        for category, count in sorted(other_hr_out_destination_counter.items(), key=lambda item: (-item[1], item[0]))
    ]
    other_hr_in_source_rows = [
        {
            "id": category,
            "category": category,
            "count": count,
        }
        for category, count in sorted(other_hr_in_source_counter.items(), key=lambda item: (-item[1], item[0]))
    ]

    return {
        "startDate": start_date.isoformat(),
        "endDate": end_date.isoformat(),
        "summary": summary,
        "summaryRows": summary_rows,
        "outflowCategoryRows": _counter_rows(outflow_counter, OUTFLOW_CATEGORY_ORDER),
        "inflowCategoryRows": _counter_rows(inflow_counter, INFLOW_CATEGORY_ORDER),
        "otherHrOutDestinationRows": other_hr_out_destination_rows,
        "otherHrInSourceRows": other_hr_in_source_rows,
        "zoneSummaryRows": zone_summary_rows,
        "leftRows": left_rows,
        "otherHrOutRows": other_hr_out_rows,
        "targetFlowRows": target_flow_rows,
        "otherHrInRows": other_hr_in_rows,
    }


def _append_table(
    worksheet,
    *,
    title: str,
    columns: Iterable[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
    start_row: int,
) -> int:
    worksheet.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    header_row = start_row + 1
    for column_index, (_, label) in enumerate(columns, start=1):
        worksheet.cell(row=header_row, column=column_index, value=label).font = Font(bold=True)

    current_row = header_row + 1
    for row in rows:
        for column_index, (key, _) in enumerate(columns, start=1):
            worksheet.cell(row=current_row, column=column_index, value=row.get(key))
        current_row += 1
    return current_row + 1


def _write_sheet(
    workbook: Workbook,
    *,
    title: str,
    columns: Iterable[tuple[str, str]],
    rows: Iterable[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title)
    headers = list(columns)
    worksheet.append([label for _, label in headers])
    for column_index in range(1, len(headers) + 1):
        worksheet.cell(row=1, column=column_index).font = Font(bold=True)
    for row in rows:
        worksheet.append([row.get(key) for key, _ in headers])

    for column_cells in worksheet.columns:
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 48)


def render_service_station_flow_workbook(
    *,
    report: dict[str, Any],
    output_path: Path,
) -> ServiceStationFlowWorkbookData:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"

    next_row = _append_table(
        summary_sheet,
        title="核心指标",
        columns=SUMMARY_EXPORT_COLUMNS,
        rows=report["summaryRows"],
        start_row=1,
    )
    next_row = _append_table(
        summary_sheet,
        title="去向分类",
        columns=(("category", "去向类型"), ("count", "人数")),
        rows=report["outflowCategoryRows"],
        start_row=next_row,
    )
    next_row = _append_table(
        summary_sheet,
        title="来源分类",
        columns=(("category", "来源类型"), ("count", "人数")),
        rows=report["inflowCategoryRows"],
        start_row=next_row,
    )
    next_row = _append_table(
        summary_sheet,
        title="转其他HR去向分布",
        columns=(("category", "期末HR子域"), ("count", "人数")),
        rows=report["otherHrOutDestinationRows"],
        start_row=next_row,
    )
    _append_table(
        summary_sheet,
        title="其他HR转入来源分布",
        columns=(("category", "期初HR子域"), ("count", "人数")),
        rows=report["otherHrInSourceRows"],
        start_row=next_row,
    )

    _write_sheet(
        workbook,
        title="战区汇总",
        columns=ZONE_EXPORT_COLUMNS,
        rows=report["zoneSummaryRows"],
    )
    _write_sheet(
        workbook,
        title="离职明细",
        columns=OUTFLOW_DETAIL_EXPORT_COLUMNS,
        rows=report["leftRows"],
    )
    _write_sheet(
        workbook,
        title="转其他HR明细",
        columns=OUTFLOW_DETAIL_EXPORT_COLUMNS,
        rows=report["otherHrOutRows"],
    )
    _write_sheet(
        workbook,
        title="目标岗位去向明细",
        columns=OUTFLOW_DETAIL_EXPORT_COLUMNS,
        rows=report["targetFlowRows"],
    )
    _write_sheet(
        workbook,
        title="其他HR转入明细",
        columns=INFLOW_DETAIL_EXPORT_COLUMNS,
        rows=report["otherHrInRows"],
    )

    workbook.save(output_path)
    return ServiceStationFlowWorkbookData(
        summary_rows=list(report["summaryRows"]),
        zone_summary_rows=list(report["zoneSummaryRows"]),
        left_rows=list(report["leftRows"]),
        other_hr_out_rows=list(report["otherHrOutRows"]),
        target_flow_rows=list(report["targetFlowRows"]),
        other_hr_in_rows=list(report["otherHrInRows"]),
    )
