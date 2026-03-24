from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from automation.reporting.low_score_feedback import display_summary_conclusion


@dataclass(frozen=True)
class AuditDistributionWorkbookData:
    batch_no: str
    assessment_version: str
    summary_rows: list[dict[str, Any]]
    detail_rows: list[dict[str, Any]]
    approval_rows: list[dict[str, Any]]
    ignored_node_names: list[str]
    document_feedback_rows: list[dict[str, Any]]


def _score_text(value: Any) -> str:
    if isinstance(value, Decimal):
        return f"{value:.1f}"
    if isinstance(value, (float, int)):
        return f"{float(value):.1f}"
    text = str(value).strip()
    return text or "<NULL>"


def _text(value: Any) -> str:
    if value is None:
        return "<NULL>"
    text = str(value).strip()
    return text or "<NULL>"


def build_document_stats(summary_rows: Iterable[dict[str, Any]]) -> dict[str, Any]:
    rows = list(summary_rows)
    conclusion_counter = Counter(_text(row.get("summary_conclusion")) for row in rows)
    score_counter = Counter(_score_text(row.get("final_score")) for row in rows)
    manual_review_documents = sum(1 for row in rows if bool(row.get("hit_manual_review")))
    low_score_documents = sum(1 for row in rows if bool(row.get("has_low_score_details")))
    return {
        "document_count": len(rows),
        "manual_review_document_count": manual_review_documents,
        "low_score_document_count": low_score_documents,
        "summary_conclusion_distribution": sorted(
            (
                {"总结论": conclusion, "单据数": count}
                for conclusion, count in conclusion_counter.items()
            ),
            key=lambda item: (-item["单据数"], item["总结论"]),
        ),
        "final_score_distribution": sorted(
            (
                {"最终信任分": score, "单据数": count}
                for score, count in score_counter.items()
            ),
            key=lambda item: (float(item["最终信任分"]), item["最终信任分"]),
        ),
    }


def build_dimension_stats(detail_rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, Any]] = {}
    for row in detail_rows:
        dimension_name = _text(row.get("dimension_name"))
        item = stats.setdefault(
            dimension_name,
            {
                "维度名称": dimension_name,
                "命中次数": 0,
                "低分命中次数": 0,
                "涉及单据数": set(),
                "低分涉及单据数": set(),
                "分值分布": Counter(),
            },
        )
        item["命中次数"] += 1
        item["涉及单据数"].add(_text(row.get("document_no")))
        score_text = _score_text(row.get("score"))
        item["分值分布"][score_text] += 1
        if bool(row.get("is_low_score")):
            item["低分命中次数"] += 1
            item["低分涉及单据数"].add(_text(row.get("document_no")))

    result: list[dict[str, Any]] = []
    for item in stats.values():
        distribution = "；".join(
            f"{score}:{count}"
            for score, count in sorted(item["分值分布"].items(), key=lambda pair: float(pair[0]))
        )
        result.append(
            {
                "维度名称": item["维度名称"],
                "命中次数": item["命中次数"],
                "低分命中次数": item["低分命中次数"],
                "涉及单据数": len(item["涉及单据数"]),
                "低分涉及单据数": len(item["低分涉及单据数"]),
                "分值分布": distribution,
            }
        )
    result.sort(key=lambda item: (-item["命中次数"], item["维度名称"]))
    return result


def build_rule_stats(detail_rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    stats: dict[tuple[str, str], dict[str, Any]] = {}
    for row in detail_rows:
        dimension_name = _text(row.get("dimension_name"))
        rule_id = _text(row.get("rule_id"))
        key = (dimension_name, rule_id)
        item = stats.setdefault(
            key,
            {
                "维度名称": dimension_name,
                "命中规则编码": rule_id,
                "命中规则说明": _text(row.get("rule_summary")),
                "命中次数": 0,
                "低分命中次数": 0,
                "涉及单据数": set(),
                "低分涉及单据数": set(),
                "最低分": None,
                "最高分": None,
            },
        )
        item["命中次数"] += 1
        item["涉及单据数"].add(_text(row.get("document_no")))
        score_value = float(row.get("score") or 0.0)
        item["最低分"] = score_value if item["最低分"] is None else min(item["最低分"], score_value)
        item["最高分"] = score_value if item["最高分"] is None else max(item["最高分"], score_value)
        if bool(row.get("is_low_score")):
            item["低分命中次数"] += 1
            item["低分涉及单据数"].add(_text(row.get("document_no")))

    result: list[dict[str, Any]] = []
    for item in stats.values():
        result.append(
            {
                "维度名称": item["维度名称"],
                "命中规则编码": item["命中规则编码"],
                "命中规则说明": item["命中规则说明"],
                "命中次数": item["命中次数"],
                "低分命中次数": item["低分命中次数"],
                "涉及单据数": len(item["涉及单据数"]),
                "低分涉及单据数": len(item["低分涉及单据数"]),
                "最低分": f"{item['最低分']:.1f}",
                "最高分": f"{item['最高分']:.1f}",
            }
        )
    result.sort(key=lambda item: (-item["命中次数"], item["维度名称"], item["命中规则编码"]))
    return result


def build_approval_node_stats(
    approval_rows: Iterable[dict[str, Any]],
    ignored_node_names: Iterable[str],
) -> list[dict[str, Any]]:
    ignored_set = {_text(name) for name in ignored_node_names}
    stats: dict[str, dict[str, Any]] = {}
    for row in approval_rows:
        node_name = _text(row.get("node_name"))
        action = _text(row.get("approval_action"))
        item = stats.setdefault(
            node_name,
            {
                "节点名称": node_name,
                "记录数": 0,
                "涉及单据数": set(),
                "审批人数": set(),
                "提交记录数": 0,
                "同意记录数": 0,
                "驳回记录数": 0,
                "待审核记录数": 0,
                "其他动作记录数": 0,
                "是否排除评分节点": "是" if node_name in ignored_set else "否",
            },
        )
        item["记录数"] += 1
        item["涉及单据数"].add(_text(row.get("document_no")))
        approver_no = _text(row.get("approver_employee_no"))
        if approver_no != "<NULL>":
            item["审批人数"].add(approver_no)
        if action == "提交":
            item["提交记录数"] += 1
        elif action == "同意":
            item["同意记录数"] += 1
        elif "驳回" in action:
            item["驳回记录数"] += 1
        elif action == "待审核":
            item["待审核记录数"] += 1
        else:
            item["其他动作记录数"] += 1

    result: list[dict[str, Any]] = []
    for item in stats.values():
        result.append(
            {
                "节点名称": item["节点名称"],
                "记录数": item["记录数"],
                "涉及单据数": len(item["涉及单据数"]),
                "审批人数": len(item["审批人数"]),
                "提交记录数": item["提交记录数"],
                "同意记录数": item["同意记录数"],
                "驳回记录数": item["驳回记录数"],
                "待审核记录数": item["待审核记录数"],
                "其他动作记录数": item["其他动作记录数"],
                "是否排除评分节点": item["是否排除评分节点"],
            }
        )
    result.sort(key=lambda item: (-item["记录数"], item["节点名称"]))
    return result


def build_document_feedback_rows(
    summary_rows: Iterable[dict[str, Any]],
    feedback_overviews: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    document_feedback_rows: list[dict[str, Any]] = []
    for row in summary_rows:
        document_no = _text(row.get("document_no"))
        if document_no == "<NULL>":
            continue
        feedback_overview = feedback_overviews.get(document_no, {})
        feedback_stats = {
            str(item.get("label") or ""): str(item.get("value") or "")
            for item in feedback_overview.get("feedbackStats", [])
            if isinstance(item, dict)
        }
        feedback_lines = [
            str(item.get("summary") or "").strip()
            for item in feedback_overview.get("feedbackGroups", [])
            if isinstance(item, dict) and str(item.get("summary") or "").strip()
        ]
        document_feedback_rows.append(
            {
                "document_no": document_no,
                "summary_conclusion_label": feedback_overview.get("summaryConclusionLabel"),
                "risk_type_count": feedback_stats.get("风险类型数", "0"),
                "affected_org_unit_count": feedback_stats.get("影响组织单位数", "0"),
                "affected_org_count": feedback_stats.get("影响组织数", "0"),
                "affected_role_count": feedback_stats.get("影响角色数", "0"),
                "raw_low_score_detail_count": feedback_stats.get("原始低分明细数", "0"),
                "feedback_summary": "\n".join(feedback_lines),
            }
        )
    return document_feedback_rows


def build_audit_distribution_workbook_data(
    *,
    batch_no: str,
    ignored_node_names: Iterable[str],
    summary_rows: Iterable[dict[str, Any]],
    detail_rows: Iterable[dict[str, Any]],
    approval_rows: Iterable[dict[str, Any]],
    feedback_overviews: dict[str, dict[str, Any]] | None = None,
) -> AuditDistributionWorkbookData:
    normalized_summary_rows = [dict(row) for row in summary_rows]
    normalized_detail_rows = [dict(row) for row in detail_rows]
    normalized_approval_rows = [dict(row) for row in approval_rows]
    normalized_ignored_node_names = [str(name).strip() for name in ignored_node_names if str(name).strip()]

    if not normalized_summary_rows:
        raise ValueError(f"No summary rows found for batch {batch_no}")

    for row in normalized_summary_rows:
        row["summary_conclusion_label"] = display_summary_conclusion(
            str(row.get("summary_conclusion") or "").strip() or None
        )

    assessment_version = str(normalized_summary_rows[0].get("assessment_version") or "").strip()
    if not assessment_version:
        raise ValueError(f"Missing assessment version for batch {batch_no}")

    document_feedback_rows = build_document_feedback_rows(
        normalized_summary_rows,
        feedback_overviews or {},
    )

    return AuditDistributionWorkbookData(
        batch_no=batch_no,
        assessment_version=assessment_version,
        summary_rows=normalized_summary_rows,
        detail_rows=normalized_detail_rows,
        approval_rows=normalized_approval_rows,
        ignored_node_names=normalized_ignored_node_names,
        document_feedback_rows=document_feedback_rows,
    )


def _fetch_latest_batch_no(cursor) -> str | None:
    cursor.execute(
        '''
        SELECT "评估批次号"
        FROM "申请单风险信任评估"
        ORDER BY "评估时间" DESC, "评估ID" DESC
        LIMIT 1
        '''
    )
    row = cursor.fetchone()
    if not row or not row[0]:
        return None
    return str(row[0]).strip() or None


def _fetch_summary_rows(cursor, batch_no: str) -> list[dict[str, object]]:
    cursor.execute(
        '''
        SELECT
            "单据编号",
            "评估版本",
            "最终信任分",
            "总结论",
            "建议动作",
            "最低命中维度",
            "最低命中角色编码",
            "最低命中组织编码",
            "是否命中人工干预",
            "是否存在低分明细",
            "低分明细条数",
            "低分明细结论"
        FROM "申请单风险信任评估"
        WHERE "评估批次号" = %s
        ORDER BY "最终信任分", "单据编号"
        ''',
        (batch_no,),
    )
    rows = cursor.fetchall()
    return [
        {
            "document_no": row[0],
            "assessment_version": row[1],
            "final_score": row[2],
            "summary_conclusion": row[3],
            "suggested_action": row[4],
            "lowest_hit_dimension": row[5],
            "lowest_hit_role_code": row[6],
            "lowest_hit_org_code": row[7],
            "hit_manual_review": row[8],
            "has_low_score_details": row[9],
            "low_score_detail_count": row[10],
            "low_score_detail_conclusion": row[11],
        }
        for row in rows
    ]


def _fetch_detail_rows(cursor, batch_no: str) -> list[dict[str, object]]:
    cursor.execute(
        '''
        SELECT
            "单据编号",
            "角色编码",
            "组织编码",
            "维度名称",
            "命中规则编码",
            "命中规则说明",
            "维度得分",
            "明细结论",
            "是否低分明细",
            "建议干预动作"
        FROM "申请单风险信任评估明细"
        WHERE "评估批次号" = %s
        ORDER BY "维度得分", "单据编号", "维度名称", "命中规则编码"
        ''',
        (batch_no,),
    )
    rows = cursor.fetchall()
    return [
        {
            "document_no": row[0],
            "role_code": row[1],
            "org_code": row[2],
            "dimension_name": row[3],
            "rule_id": row[4],
            "rule_summary": row[5],
            "score": row[6],
            "detail_conclusion": row[7],
            "is_low_score": row[8],
            "intervention_action": row[9],
        }
        for row in rows
    ]


def _fetch_approval_rows(cursor, document_nos: list[str]) -> list[dict[str, object]]:
    if not document_nos:
        return []
    cursor.execute(
        '''
        SELECT
            "单据编号",
            "节点名称",
            "审批动作",
            "工号"
        FROM "申请单审批记录"
        WHERE "单据编号" = ANY(%s)
        ORDER BY "单据编号", "审批记录顺序号"
        ''',
        (document_nos,),
    )
    rows = cursor.fetchall()
    return [
        {
            "document_no": row[0],
            "node_name": row[1],
            "approval_action": row[2],
            "approver_employee_no": row[3],
        }
        for row in rows
    ]


def load_audit_distribution_workbook_data(
    *,
    store: Any,
    batch_no: str,
    ignored_node_names: Iterable[str],
) -> AuditDistributionWorkbookData:
    normalized_batch_no = str(batch_no or "").strip()
    with store.connect() as connection:
        with connection.cursor() as cursor:
            resolved_batch_no = normalized_batch_no or _fetch_latest_batch_no(cursor)
            if not resolved_batch_no:
                raise ValueError("No assessment batch found in 申请单风险信任评估")
            summary_rows = _fetch_summary_rows(cursor, resolved_batch_no)
            detail_rows = _fetch_detail_rows(cursor, resolved_batch_no)
            document_nos = [
                str(row.get("document_no") or "").strip()
                for row in summary_rows
                if str(row.get("document_no") or "").strip()
            ]
            approval_rows = _fetch_approval_rows(cursor, document_nos)

    feedback_overviews = store.fetch_document_feedback_overviews(
        assessment_batch_no=resolved_batch_no,
        document_nos=[
            str(row.get("document_no") or "").strip()
            for row in summary_rows
            if str(row.get("document_no") or "").strip()
        ],
    )
    return build_audit_distribution_workbook_data(
        batch_no=resolved_batch_no,
        ignored_node_names=ignored_node_names,
        summary_rows=summary_rows,
        detail_rows=detail_rows,
        approval_rows=approval_rows,
        feedback_overviews=feedback_overviews,
    )


def export_audit_distribution_workbook(
    *,
    store: Any,
    batch_no: str,
    ignored_node_names: Iterable[str],
    output_path: Path,
) -> AuditDistributionWorkbookData:
    workbook_data = load_audit_distribution_workbook_data(
        store=store,
        batch_no=batch_no,
        ignored_node_names=ignored_node_names,
    )
    render_audit_distribution_workbook(
        batch_no=workbook_data.batch_no,
        assessment_version=workbook_data.assessment_version,
        summary_rows=workbook_data.summary_rows,
        detail_rows=workbook_data.detail_rows,
        approval_rows=workbook_data.approval_rows,
        ignored_node_names=workbook_data.ignored_node_names,
        document_feedback_rows=workbook_data.document_feedback_rows,
        output_path=output_path,
    )
    return workbook_data


def render_audit_distribution_workbook(
    *,
    batch_no: str,
    assessment_version: str,
    summary_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    approval_rows: list[dict[str, Any]],
    ignored_node_names: list[str],
    document_feedback_rows: list[dict[str, Any]] | None = None,
    output_path: Path,
) -> Path:
    workbook = Workbook()
    document_stats = build_document_stats(summary_rows)
    dimension_stats = build_dimension_stats(detail_rows)
    rule_stats = build_rule_stats(detail_rows)
    approval_node_stats = build_approval_node_stats(approval_rows, ignored_node_names)
    low_score_rows = [row for row in detail_rows if bool(row.get("is_low_score"))]

    overview_sheet = workbook.active
    overview_sheet.title = "批次总览"
    _append_section(
        overview_sheet,
        "批次信息",
        ["指标", "数值"],
        [
            ["评估批次号", batch_no],
            ["评估版本", assessment_version],
            ["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["单据数", document_stats["document_count"]],
            ["明细行数", len(detail_rows)],
            ["审批记录数", len(approval_rows)],
            ["命中人工干预单据数", document_stats["manual_review_document_count"]],
            ["存在低分明细单据数", document_stats["low_score_document_count"]],
        ],
    )
    _append_section(
        overview_sheet,
        "总结论分布",
        ["总结论", "单据数"],
        [[row["总结论"], row["单据数"]] for row in document_stats["summary_conclusion_distribution"]],
    )
    _append_section(
        overview_sheet,
        "最终信任分分布",
        ["最终信任分", "单据数"],
        [[row["最终信任分"], row["单据数"]] for row in document_stats["final_score_distribution"]],
    )

    _append_table_sheet(
        workbook,
        "单据结果",
        [
            "单据编号",
            "最终信任分",
            "展示结论",
            "总结论",
            "建议动作",
            "最低命中维度",
            "最低命中角色编码",
            "最低命中组织编码",
            "是否命中人工干预",
            "低分明细条数",
            "低分明细结论",
        ],
        [
            [
                _text(row.get("document_no")),
                _score_text(row.get("final_score")),
                _text(row.get("summary_conclusion_label")),
                _text(row.get("summary_conclusion")),
                _text(row.get("suggested_action")),
                _text(row.get("lowest_hit_dimension")),
                _text(row.get("lowest_hit_role_code")),
                _text(row.get("lowest_hit_org_code")),
                "是" if bool(row.get("hit_manual_review")) else "否",
                int(row.get("low_score_detail_count") or 0),
                _text(row.get("low_score_detail_conclusion")),
            ]
            for row in sorted(
                summary_rows,
                key=lambda item: (float(item.get("final_score") or 0.0), _text(item.get("document_no"))),
            )
        ],
    )
    if document_feedback_rows:
        _append_table_sheet(
            workbook,
            "104聚合反馈",
            [
                "单据编号",
                "展示结论",
                "风险类型数",
                "影响组织单位数",
                "影响组织数",
                "影响角色数",
                "原始低分明细数",
                "聚合反馈摘要",
            ],
            [
                [
                    _text(row.get("document_no")),
                    _text(row.get("summary_conclusion_label")),
                    _text(row.get("risk_type_count")),
                    _text(row.get("affected_org_unit_count")),
                    _text(row.get("affected_org_count")),
                    _text(row.get("affected_role_count")),
                    _text(row.get("raw_low_score_detail_count")),
                    _text(row.get("feedback_summary")),
                ]
                for row in sorted(document_feedback_rows, key=lambda item: _text(item.get("document_no")))
            ],
        )
    _append_table_sheet(
        workbook,
        "维度分布",
        ["维度名称", "命中次数", "低分命中次数", "涉及单据数", "低分涉及单据数", "分值分布"],
        [
            [
                row["维度名称"],
                row["命中次数"],
                row["低分命中次数"],
                row["涉及单据数"],
                row["低分涉及单据数"],
                row["分值分布"],
            ]
            for row in dimension_stats
        ],
    )
    _append_table_sheet(
        workbook,
        "规则分布",
        ["维度名称", "命中规则编码", "命中规则说明", "命中次数", "低分命中次数", "涉及单据数", "低分涉及单据数", "最低分", "最高分"],
        [
            [
                row["维度名称"],
                row["命中规则编码"],
                row["命中规则说明"],
                row["命中次数"],
                row["低分命中次数"],
                row["涉及单据数"],
                row["低分涉及单据数"],
                row["最低分"],
                row["最高分"],
            ]
            for row in rule_stats
        ],
    )
    _append_table_sheet(
        workbook,
        "审批节点分布",
        ["节点名称", "记录数", "涉及单据数", "审批人数", "提交记录数", "同意记录数", "驳回记录数", "待审核记录数", "其他动作记录数", "是否排除评分节点"],
        [
            [
                row["节点名称"],
                row["记录数"],
                row["涉及单据数"],
                row["审批人数"],
                row["提交记录数"],
                row["同意记录数"],
                row["驳回记录数"],
                row["待审核记录数"],
                row["其他动作记录数"],
                row["是否排除评分节点"],
            ]
            for row in approval_node_stats
        ],
    )
    _append_table_sheet(
        workbook,
        "低分明细",
        ["单据编号", "维度名称", "命中规则编码", "维度得分", "角色编码", "组织编码", "建议干预动作", "明细结论"],
        [
            [
                _text(row.get("document_no")),
                _text(row.get("dimension_name")),
                _text(row.get("rule_id")),
                _score_text(row.get("score")),
                _text(row.get("role_code")),
                _text(row.get("org_code")),
                _text(row.get("intervention_action")),
                _text(row.get("detail_conclusion")),
            ]
            for row in sorted(
                low_score_rows,
                key=lambda item: (
                    float(item.get("score") or 0.0),
                    _text(item.get("document_no")),
                    _text(item.get("dimension_name")),
                    _text(item.get("rule_id")),
                ),
            )
        ],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _append_section(sheet, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    row_index = sheet.max_row + 1 if sheet.max_row > 1 or sheet["A1"].value else 1
    sheet.cell(row=row_index, column=1, value=title)
    sheet.cell(row=row_index, column=1).font = Font(bold=True)
    row_index += 1
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=header)
        cell.font = Font(bold=True)
    for row in rows:
        row_index += 1
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
    row_index += 2
    _autosize_columns(sheet)


def _append_table_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)
    for row_index, row in enumerate(rows, start=2):
        for col_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=col_index, value=value)
    sheet.freeze_panes = "A2"
    _autosize_columns(sheet)


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(text), 80))
        sheet.column_dimensions[column_letter].width = max(max_length + 2, 12)
