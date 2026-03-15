from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

ROSTER_FIELD_SPECS = [
    ("序号", "serial_no", "TEXT"),
    ("公司", "company_name", "TEXT"),
    ("公司ID", "company_id", "TEXT"),
    ("部门", "department_name", "TEXT"),
    ("部门长文本", "department_long_text", "TEXT"),
    ("部门ID", "department_id", "TEXT"),
    ("部门所属城市", "department_city", "TEXT"),
    ("人员编号", "employee_no", "TEXT"),
    ("姓名", "employee_name", "TEXT"),
    ("职位编码", "position_code", "TEXT"),
    ("职位名称", "position_name", "TEXT"),
    ("具体岗位(新)", "specific_post_name", "TEXT"),
    ("是否关键岗位", "critical_post_flag", "TEXT"),
    ("岗位族", "post_family", "TEXT"),
    ("职务", "job_title", "TEXT"),
    ("一级职能名称", "level1_function_name", "TEXT"),
    ("二级职能名称", "level2_function_name", "TEXT"),
    ("标准岗位编码", "standard_position_code", "TEXT"),
    ("标准岗位名称", "standard_position_name", "TEXT"),
    ("上级编码", "superior_code", "TEXT"),
    ("上级名称", "superior_name", "TEXT"),
    ("对接HR工号", "hr_partner_employee_no", "TEXT"),
    ("对接HR姓名", "hr_partner_name", "TEXT"),
    ("班长", "team_leader_name", "TEXT"),
    ("员工组", "employee_group", "TEXT"),
    ("员工子组", "employee_subgroup", "TEXT"),
    ("子标签", "sub_label", "TEXT"),
    ("入职日期", "entry_date", "DATE"),
    ("司龄", "company_seniority_text", "TEXT"),
    ("性别", "gender", "TEXT"),
    ("出生日期", "birth_date", "DATE"),
    ("生日类型", "birthday_type", "TEXT"),
    ("实际生日", "actual_birthday", "TEXT"),
    ("年龄", "age_text", "TEXT"),
    ("户口所在省", "household_registration_province", "TEXT"),
    ("户口所在市", "household_registration_city", "TEXT"),
    ("户籍地", "registered_residence", "TEXT"),
    ("参保地", "social_security_location", "TEXT"),
    ("婚姻状况", "marital_status", "TEXT"),
    ("学历", "education_level", "TEXT"),
    ("学习形式", "study_mode", "TEXT"),
    ("毕业时间", "graduation_date", "DATE"),
    ("学位", "degree", "TEXT"),
    ("学校", "school_name", "TEXT"),
    ("其他院校", "alternate_school_name", "TEXT"),
    ("院校专业", "major_name", "TEXT"),
    ("证件号码", "id_number", "TEXT"),
    ("联系电话", "phone_number", "TEXT"),
    ("万科邮箱", "vanke_email", "TEXT"),
    ("万物云邮箱", "onewo_email", "TEXT"),
    ("域账号", "domain_account", "TEXT"),
    ("雇佣状态", "employment_status", "TEXT"),
    ("部门分类", "department_category", "TEXT"),
    ("部门子分类", "department_subcategory", "TEXT"),
    ("是否外盘人员", "external_roster_flag", "TEXT"),
    ("合同类型", "contract_type", "TEXT"),
    ("合同签订主体", "contract_signing_entity", "TEXT"),
    ("合同开始日期", "contract_start_date", "DATE"),
    ("合同结束日期", "contract_end_date", "DATE"),
    ("参加工作时间", "start_work_date", "DATE"),
    ("服务万科时间", "service_start_date_at_vanke", "DATE"),
    ("最新入职万科日期", "latest_entry_date_to_vanke", "DATE"),
    ("司龄起算时间", "seniority_start_date", "DATE"),
    ("工龄", "working_years_text", "TEXT"),
    ("储备见习类型", "trainee_type", "TEXT"),
    ("储备见习开始日期", "trainee_start_date", "DATE"),
    ("储备见习结束日期", "trainee_end_date", "DATE"),
    ("储备岗位名称", "trainee_post_name", "TEXT"),
    ("岗位备注（物业）", "post_remark_property", "TEXT"),
    ("序列属性", "sequence_attribute", "TEXT"),
    ("序列类型", "sequence_type", "TEXT"),
    ("组织路径ID", "org_path_id", "TEXT"),
    ("组织路径名称", "org_path_name", "TEXT"),
    ("合同期限", "contract_term", "TEXT"),
    ("是否服过兵役", "military_service_flag", "TEXT"),
    ("退役证编号", "discharge_certificate_no", "TEXT"),
    ("退伍证类型", "discharge_certificate_type", "TEXT"),
    ("入伍时间", "enlistment_date", "DATE"),
    ("退役时间", "discharge_date", "DATE"),
    ("领导力类别", "leadership_category", "TEXT"),
    ("证件是否永久有效", "id_document_permanent_flag", "TEXT"),
    ("证件有效期起", "id_document_valid_from", "DATE"),
    ("证件有效期止", "id_document_valid_to", "DATE"),
    ("是否主任职", "primary_position_flag", "TEXT"),
    ("任职类型", "employment_type", "TEXT"),
    ("民族", "ethnicity", "TEXT"),
    ("户口性质", "household_registration_type", "TEXT"),
    ("户口所在地", "household_registration_location", "TEXT"),
    ("出生地", "birthplace", "TEXT"),
    ("籍贯", "native_place", "TEXT"),
    ("国籍", "nationality", "TEXT"),
    ("出生国家", "birth_country", "TEXT"),
    ("出生地所在省", "birthplace_province", "TEXT"),
    ("出生地所在市", "birthplace_city", "TEXT"),
    ("校招届别", "campus_recruitment_cohort", "TEXT"),
    ("政治面貌", "political_status", "TEXT"),
    ("所属党组织全称", "party_organization_name", "TEXT"),
    ("党员状态", "party_member_status", "TEXT"),
]

HEADER_ALIASES = {
    "具体岗位（新）": "specific_post_name",
    "岗位备注(物业)": "post_remark_property",
}

HEADER_MAP = {header: field_name for header, field_name, _ in ROSTER_FIELD_SPECS}
HEADER_MAP.update(HEADER_ALIASES)
HEADER_CANDIDATES = ["公司", "公司ID", "部门", "部门ID", "人员编号", "姓名", "入职日期"]
ROSTER_FIELD_NAMES = [field_name for _, field_name, _ in ROSTER_FIELD_SPECS]
ROSTER_DATE_FIELDS = {field_name for _, field_name, column_type in ROSTER_FIELD_SPECS if column_type == "DATE"}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _rows_from_sheet(sheet) -> list[list[str]]:
    return [[normalize_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def _load_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError(
            "Missing dependency: openpyxl. Run `.venv/bin/python -m pip install -r automation/requirements.txt`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        if getattr(sheet, "max_row", 0) == 1 and getattr(sheet, "max_column", 0) == 1:
            try:
                sheet.reset_dimensions()
            except Exception:  # noqa: BLE001
                pass
        rows = _rows_from_sheet(sheet)
    finally:
        workbook.close()

    if len(rows) > 1:
        return rows

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return _rows_from_sheet(sheet)
    finally:
        workbook.close()


def _load_xls_rows(path: Path) -> list[list[str]]:
    try:
        import xlrd
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError(
            "Missing dependency: xlrd. Run `.venv/bin/python -m pip install -r automation/requirements.txt`."
        ) from exc

    book = xlrd.open_workbook(path)
    sheet = next((book.sheet_by_index(idx) for idx in range(book.nsheets) if book.sheet_by_index(idx).nrows > 0), book.sheet_by_index(0))
    rows: list[list[str]] = []
    for row_idx in range(sheet.nrows):
        current: list[str] = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                parts = xlrd.xldate_as_tuple(cell.value, book.datemode)
                if parts[3:] == (0, 0, 0):
                    value = date(parts[0], parts[1], parts[2])
                else:
                    value = datetime(*parts)
            else:
                value = cell.value
            current.append(normalize_text(value))
        rows.append(current)
    return rows


def _load_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(path)
    if suffix == ".xls":
        return _load_xls_rows(path)
    raise ValueError(f"Unsupported roster file type: {path.suffix}")


def _extract_query_date(rows: list[list[str]]) -> date | None:
    for row in rows[:20]:
        for cell in row:
            if "查询日期" not in cell:
                continue
            matched = re.search(r"查询日期[:：]\s*(\d{4}-\d{2}-\d{2})", cell)
            if matched:
                return date.fromisoformat(matched.group(1))
    return None


def _find_header_row(rows: list[list[str]]) -> int:
    best_index = -1
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        score = sum(1 for cell in row if normalize_text(cell) in HEADER_CANDIDATES)
        if score > best_score:
            best_score = score
            best_index = idx
        if score >= 5:
            return idx
    if best_index < 0 or best_score < 3:
        raise ValueError("Unable to identify roster header row from workbook")
    return best_index


def _normalize_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for idx, value in enumerate(row):
        header = normalize_text(value)
        if not header:
            header = f"未命名列{idx + 1}"
        counts[header] = counts.get(header, 0) + 1
        if counts[header] > 1:
            header = f"{header}_{counts[header]}"
        headers.append(header)
    return headers


def _find_unmapped_headers(headers: list[str]) -> list[str]:
    unmapped: list[str] = []
    for header in headers:
        normalized = normalize_text(header)
        if not normalized or normalized.startswith("未命名列"):
            continue
        if normalized in HEADER_MAP:
            continue
        unmapped.append(normalized)
    return unmapped


def _extract_extra_columns(record: dict[str, str]) -> dict[str, str]:
    extra_columns: dict[str, str] = {}
    for header, value in record.items():
        normalized = normalize_text(header)
        if not normalized or normalized.startswith("未命名列"):
            continue
        if normalized in HEADER_MAP:
            continue
        cell_value = normalize_text(value)
        if not cell_value:
            continue
        extra_columns[normalized] = cell_value
    return extra_columns


def _is_empty_row(row: list[str]) -> bool:
    return all(not normalize_text(cell) for cell in row)


def _is_noise_row(row: list[str], headers: list[str]) -> bool:
    normalized = [normalize_text(cell) for cell in row]
    if _is_empty_row(normalized):
        return True
    if normalized[: len(headers)] == headers[: len(normalized)]:
        return True
    joined = " ".join(cell for cell in normalized if cell)
    if not joined:
        return True
    if any(key in joined for key in ["查询日期：", "导出时间", "在职人员花名册"]):
        return True
    if len([cell for cell in normalized if cell]) < 2:
        return True
    return False


def _parse_row(headers: list[str], row: list[str], excel_row_index: int) -> dict[str, Any] | None:
    normalized_row = [normalize_text(cell) for cell in row]
    record = {headers[idx]: normalized_row[idx] if idx < len(normalized_row) else "" for idx in range(len(headers))}
    standard: dict[str, Any] = {
        "row_no": excel_row_index,
        "extra_columns": _extract_extra_columns(record),
    }
    for field_name in ROSTER_FIELD_NAMES:
        standard[field_name] = ""
    for header, field_name in HEADER_MAP.items():
        if header in record and not standard[field_name]:
            standard[field_name] = record.get(header, "")
    if not standard.get("employee_no"):
        return None
    return standard


def parse_roster_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path)
    rows = _load_rows(workbook_path)
    if not rows:
        raise ValueError(f"Workbook is empty: {workbook_path}")

    query_date = _extract_query_date(rows)
    header_row_index = _find_header_row(rows)
    headers = _normalize_headers(rows[header_row_index])
    unmapped_headers = _find_unmapped_headers(headers)

    records: list[dict[str, Any]] = []
    for offset, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        if _is_noise_row(row, headers):
            continue
        parsed = _parse_row(headers, row, offset)
        if parsed is None:
            continue
        records.append(parsed)

    if not records:
        raise ValueError(f"No roster data rows found after header parsing: {workbook_path}")

    return {
        "file_path": str(workbook_path),
        "file_name": workbook_path.name,
        "query_date": query_date,
        "headers": headers,
        "unmapped_headers": unmapped_headers,
        "records": records,
        "row_count": len(records),
    }
