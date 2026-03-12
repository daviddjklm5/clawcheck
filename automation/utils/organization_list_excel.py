from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

ORGLIST_FIELD_SPECS = [
    ("序号", "row_no"),
    ("行政组织编码", "org_code"),
    ("行政组织名称", "org_name"),
    ("行政组织类型", "org_type"),
    ("上级行政组织", "parent_org_name"),
    ("上级行政组织编码", "parent_org_code"),
    ("所属公司", "company_name"),
    ("行政组织层级", "org_level"),
    ("所在城市", "city_name"),
    ("物理层级", "physical_level"),
    ("部门子分类.编码", "dept_subcategory_code"),
    ("部门子分类.名称", "dept_subcategory_name"),
    ("部门分类.编码", "dept_category_code"),
    ("部门分类.名称", "dept_category_name"),
    ("组织长名称", "org_full_name"),
    ("组织负责人", "org_manager_name"),
    ("责任HR工号", "hr_owner_employee_no"),
    ("责任HR姓名", "hr_owner_name"),
    ("责任HR含下级组织", "hr_owner_include_children_flag"),
    ("责任HR是否透出", "hr_owner_exposed_flag"),
]

HEADER_MAP = {header: field_name for header, field_name in ORGLIST_FIELD_SPECS}
HEADER_CANDIDATES = [
    "序号",
    "行政组织编码",
    "行政组织名称",
    "行政组织类型",
    "上级行政组织",
    "所属公司",
    "组织长名称",
]
ORGLIST_FIELD_NAMES = [field_name for _, field_name in ORGLIST_FIELD_SPECS]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.date().isoformat()
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    text = str(value).replace("\r", " ").replace("\n", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _rows_from_sheet(sheet) -> list[list[str]]:
    return [[normalize_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def _load_xlsx_rows(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("Missing dependency: openpyxl. Run `pip install -r automation/requirements.txt`.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        if getattr(sheet, "max_row", 0) == 1 and getattr(sheet, "max_column", 0) == 1:
            try:
                sheet.reset_dimensions()
            except Exception:  # noqa: BLE001
                pass
        rows = _rows_from_sheet(sheet)
    finally:
        workbook.close()

    if len(rows) > 1 and any(len(row) > 1 for row in rows[:3]):
        return rows

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        return _rows_from_sheet(sheet)
    finally:
        workbook.close()


def _load_xls_rows(path: Path) -> list[list[str]]:
    try:
        import xlrd
    except ModuleNotFoundError as exc:  # pragma: no cover
        raise ModuleNotFoundError("Missing dependency: xlrd. Run `pip install -r automation/requirements.txt`.") from exc

    book = xlrd.open_workbook(path)
    sheet = next((book.sheet_by_index(idx) for idx in range(book.nsheets) if book.sheet_by_index(idx).nrows > 0), book.sheet_by_index(0))
    rows: list[list[str]] = []
    for row_idx in range(sheet.nrows):
        current: list[str] = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            if cell.ctype == xlrd.XL_CELL_DATE:
                parts = xlrd.xldate_as_tuple(cell.value, book.datemode)
                if parts[3:] == (0, 0, 0):
                    value = date(parts[0], parts[1], parts[2])
                else:
                    value = datetime(*parts)
            else:
                value = cell.value
            current.append(normalize_text(value))
        rows.append(current)
    return rows


def _load_rows(path: Path) -> list[list[str]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_rows(path)
    if suffix == ".xls":
        return _load_xls_rows(path)
    raise ValueError(f"Unsupported organization list file type: {path.suffix}")


def _find_header_row(rows: list[list[str]]) -> int:
    best_index = -1
    best_score = -1
    for idx, row in enumerate(rows[:20]):
        score = sum(1 for cell in row if normalize_text(cell) in HEADER_CANDIDATES)
        if score > best_score:
            best_score = score
            best_index = idx
        if score >= 5:
            return idx
    if best_index < 0 or best_score < 3:
        raise ValueError("Unable to identify organization list header row from workbook")
    return best_index


def _normalize_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for idx, value in enumerate(row):
        header = normalize_text(value)
        if not header:
            header = f"未命名列{idx + 1}"
        counts[header] = counts.get(header, 0) + 1
        if counts[header] > 1:
            header = f"{header}_{counts[header]}"
        headers.append(header)
    return headers


def _find_unmapped_headers(headers: list[str]) -> list[str]:
    unmapped: list[str] = []
    for header in headers:
        normalized = normalize_text(header)
        if not normalized or normalized.startswith("未命名列"):
            continue
        if normalized in HEADER_MAP:
            continue
        unmapped.append(normalized)
    return unmapped


def _extract_extra_columns(record: dict[str, str]) -> dict[str, str]:
    extra_columns: dict[str, str] = {}
    for header, value in record.items():
        normalized = normalize_text(header)
        if not normalized or normalized.startswith("未命名列"):
            continue
        if normalized in HEADER_MAP:
            continue
        cell_value = normalize_text(value)
        if not cell_value:
            continue
        extra_columns[normalized] = cell_value
    return extra_columns


def _is_empty_row(row: list[str]) -> bool:
    return all(not normalize_text(cell) for cell in row)


def _is_noise_row(row: list[str], headers: list[str]) -> bool:
    normalized = [normalize_text(cell) for cell in row]
    if _is_empty_row(normalized):
        return True
    if normalized[: len(headers)] == headers[: len(normalized)]:
        return True
    joined = " ".join(cell for cell in normalized if cell)
    if not joined:
        return True
    return len([cell for cell in normalized if cell]) < 2


def _parse_row(headers: list[str], row: list[str]) -> dict[str, Any] | None:
    normalized_row = [normalize_text(cell) for cell in row]
    record = {headers[idx]: normalized_row[idx] if idx < len(normalized_row) else "" for idx in range(len(headers))}
    standard: dict[str, Any] = {"extra_columns": _extract_extra_columns(record)}
    for field_name in ORGLIST_FIELD_NAMES:
        standard[field_name] = ""
    for header, field_name in HEADER_MAP.items():
        standard[field_name] = record.get(header, "")
    if not standard.get("org_code"):
        return None
    return standard


def parse_organization_list_workbook(path: str | Path) -> dict[str, Any]:
    workbook_path = Path(path)
    rows = _load_rows(workbook_path)
    if not rows:
        raise ValueError(f"Workbook is empty: {workbook_path}")

    header_row_index = _find_header_row(rows)
    headers = _normalize_headers(rows[header_row_index])
    unmapped_headers = _find_unmapped_headers(headers)

    records: list[dict[str, Any]] = []
    for row in rows[header_row_index + 1 :]:
        if _is_noise_row(row, headers):
            continue
        parsed = _parse_row(headers, row)
        if parsed is None:
            continue
        records.append(parsed)

    if not records:
        raise ValueError(f"No organization list rows found after header parsing: {workbook_path}")

    return {
        "file_path": str(workbook_path),
        "file_name": workbook_path.name,
        "headers": headers,
        "unmapped_headers": unmapped_headers,
        "records": records,
        "row_count": len(records),
    }
